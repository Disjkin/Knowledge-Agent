"""
JSONL/JSON 测试结果 → Excel 转换脚本
用法: uv run python json_to_excel.py [input.json] [output.xlsx]
"""
import json
import sys
from pathlib import Path

# 修复 Windows GBK 编码
if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if sys.stderr and hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要 openpyxl: pip install openpyxl")
    sys.exit(1)

# ── 样式 ──
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
BODY_FONT = Font(name="微软雅黑", size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# 状态颜色
STATUS_FILLS = {
    "ok": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "error": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "timeout": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "exception": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "http_error": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}


def classify(qid: str) -> str:
    """根据 qid 前缀返回中文类别名。"""
    mapping = {
        "A": "知识库命中",
        "B": "时事新闻",
        "C": "科技/AI",
        "D": "生活/健康",
        "E": "英文查询",
        "F": "财经/商业",
        "G": "教育/考试",
        "H": "旅游/文化",
        "I": "边界异常",
    }
    # v3 把网搜题统一用 W 前缀，按编号区间细分到上述类别（v2 的 B-I 不受影响）
    if qid.startswith("W"):
        try:
            n = int(qid[1:])
        except ValueError:
            return "网搜兜底"
        if 1 <= n <= 16: return "时事新闻"
        if 17 <= n <= 28: return "科技/AI"
        if 29 <= n <= 40: return "生活/健康"
        if 41 <= n <= 50: return "英文查询"
        if 51 <= n <= 60: return "财经/商业"
        if 61 <= n <= 68: return "教育/考试"
        if 69 <= n <= 76: return "旅游/文化"
        if 77 <= n <= 80: return "历史/科学"
        return "网搜兜底"
    return mapping.get(qid[0], "其他")


def build_summary_sheet(ws, summary: dict):
    """Sheet 1: 汇总概览。"""
    ws.title = "汇总概览"
    ws.sheet_properties.tabColor = "4472C4"

    # 兼容不同版本 summary（v3 把 exception 并入 error，可能缺键）-> 补默认值
    s = {"total": 0, "ok": 0, "error": 0, "timeout": 0, "exception": 0,
         "web_search_used": 0, "kb_only": 0, "avg_duration_ms": 0, "avg_answer_chars": 0}
    s.update(summary)
    summary = s

    rows = [
        ("指标", "数值", "占比"),
        ("总 Query 数", summary["total"], ""),
        ("成功 (ok)", summary["ok"], f"{summary['ok']/summary['total']*100:.0f}%") if summary["total"] else ("成功 (ok)", 0, ""),
        ("错误 (error)", summary["error"], f"{summary['error']/summary['total']*100:.0f}%") if summary["total"] else ("错误 (error)", 0, ""),
        ("超时 (timeout)", summary["timeout"], f"{summary['timeout']/summary['total']*100:.0f}%") if summary["total"] else ("超时 (timeout)", 0, ""),
        ("异常 (exception)", summary["exception"], f"{summary['exception']/summary['total']*100:.0f}%") if summary["total"] else ("异常 (exception)", 0, ""),
        ("触发网搜兜底", summary["web_search_used"], f"{summary['web_search_used']/summary['total']*100:.0f}%") if summary["total"] else ("触发网搜兜底", 0, ""),
        ("仅本地 KB", summary["kb_only"], f"{summary['kb_only']/summary['total']*100:.0f}%") if summary["total"] else ("仅本地 KB", 0, ""),
        ("平均响应耗时(ms)", summary["avg_duration_ms"], ""),
        ("平均答案长度(字)", summary["avg_answer_chars"], ""),
    ]

    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
            if r == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER_ALIGN
            else:
                cell.font = BODY_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10


def build_detail_sheet(ws, results: list):
    """Sheet 2: 逐条明细。"""
    ws.title = "逐条明细"
    ws.sheet_properties.tabColor = "70AD47"

    headers = [
        "ID", "类别", "Query", "状态", "网搜",
        "KB来源数", "Web来源数", "耗时(ms)", "答案字数",
        "Trace ID", "答案预览", "错误信息",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for r, item in enumerate(results, 2):
        qid = item.get("qid", "")
        row_data = [
            qid,
            classify(qid),
            item.get("query", ""),
            item.get("status", ""),
            "✓" if item.get("used_web_search") else "",
            item.get("sources_kb", 0),
            item.get("sources_web", 0),
            round(item.get("duration_ms", 0)),
            item.get("answer_chars", 0),
            item.get("trace_id", ""),
            item.get("answer_preview", ""),
            item.get("error") or "",
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            # 状态列着色
            if c == 4:  # 状态
                fill = STATUS_FILLS.get(str(val))
                if fill:
                    cell.fill = fill
                cell.alignment = CENTER_ALIGN
            elif c in (1, 2, 5, 6, 7, 8, 9):  # 短列居中
                cell.alignment = CENTER_ALIGN
            else:  # 长文本换行
                cell.alignment = WRAP_ALIGN

    # 列宽
    widths = [6, 12, 36, 10, 6, 9, 10, 10, 9, 34, 50, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"
    # 自动筛选
    ws.auto_filter.ref = f"A1:L{len(results)+1}"


def build_category_sheet(ws, results: list):
    """Sheet 3: 分类统计。"""
    ws.title = "分类统计"
    ws.sheet_properties.tabColor = "FFC000"

    from collections import defaultdict
    cats = defaultdict(list)
    for r in results:
        cats[classify(r.get("qid", ""))].append(r)

    headers = ["类别", "总数", "成功", "网搜", "仅KB", "错误", "超时", "平均耗时(ms)", "平均字数"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    cat_order = ["知识库命中", "时事新闻", "科技/AI", "生活/健康", "英文查询", "财经/商业", "教育/考试", "旅游/文化", "历史/科学", "边界异常"]
    for r, cat_name in enumerate(cat_order, 2):
        items = cats.get(cat_name, [])
        if not items:
            continue
        total = len(items)
        ok = sum(1 for x in items if x["status"] == "ok")
        web = sum(1 for x in items if x.get("used_web_search"))
        kb_only = sum(1 for x in items if x["status"] == "ok" and not x.get("used_web_search"))
        err = sum(1 for x in items if x["status"] in ("error", "exception", "http_error"))
        timeout = sum(1 for x in items if x["status"] == "timeout")
        avg_ms = round(sum(x.get("duration_ms", 0) for x in items) / total)
        avg_chars = round(sum(x.get("answer_chars", 0) for x in items if x["status"] == "ok") / max(ok, 1))

        row_data = [cat_name, total, ok, web, kb_only, err, timeout, avg_ms, avg_chars]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

    # 合计行
    total_row = len(cat_order) + 2
    total = len(results)
    ok = sum(1 for x in results if x["status"] == "ok")
    web = sum(1 for x in results if x.get("used_web_search"))
    kb_only = sum(1 for x in results if x["status"] == "ok" and not x.get("used_web_search"))
    err = sum(1 for x in results if x["status"] in ("error", "exception", "http_error"))
    timeout = sum(1 for x in results if x["status"] == "timeout")
    avg_ms = round(sum(x.get("duration_ms", 0) for x in results) / total)
    avg_chars = round(sum(x.get("answer_chars", 0) for x in results if x["status"] == "ok") / max(ok, 1))

    row_data = ["合计", total, ok, web, kb_only, err, timeout, avg_ms, avg_chars]
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=total_row, column=c, value=val)
        cell.font = Font(name="微软雅黑", bold=True, size=10)
        cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    widths = [14, 8, 8, 8, 8, 8, 8, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def main():
    input_file = sys.argv[1] if len(sys.argv) > 1 else "test_100_queries_result_v2.json"
    output_file = sys.argv[2] if len(sys.argv) > 2 else Path(input_file).stem + ".xlsx"

    with open(input_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()

    # Sheet 1: 汇总概览
    ws1 = wb.active
    build_summary_sheet(ws1, data.get("summary", {}))

    # Sheet 2: 逐条明细
    ws2 = wb.create_sheet()
    build_detail_sheet(ws2, data.get("results", []))

    # Sheet 3: 分类统计
    ws3 = wb.create_sheet()
    build_category_sheet(ws3, data.get("results", []))

    wb.save(output_file)
    print(f"✅ 已生成 Excel: {output_file}")
    print(f"   包含 3 个 Sheet: 汇总概览 / 逐条明细 / 分类统计")
    print(f"   总数据量: {len(data.get('results', []))} 条")


if __name__ == "__main__":
    main()
